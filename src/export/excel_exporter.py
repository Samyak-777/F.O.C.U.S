"""
Excel attendance export using openpyxl.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import hashlib
import io
from config.constants import AttendanceStatus


def export_attendance_excel(session_data: dict, records: list, overrides: list) -> bytes:
    """Generate styled Excel attendance export."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title rows
    ws.append(["FOCUS Attendance Report"])
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(name="Calibri", bold=True, size=16)

    ws.append([
        f"Batch: {session_data.get('batch_id', '')}",
        f"Date: {session_data.get('date', '')}",
        f"Faculty: {session_data.get('faculty_name', '')}"
    ])
    ws.append([f"Generated: {datetime.utcnow().isoformat()} UTC"])
    ws.append([])  # Empty row

    # Headers
    headers = ["Roll No.", "Name", "Status", "AI Confidence",
               "Override", "Override By", "Comment"]
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    override_map = {o["roll_number"]: o for o in overrides}
    for rec in records:
        roll = rec["roll_number"]
        override = override_map.get(roll)

        if rec["status"] == AttendanceStatus.CONSENT_WITHDRAWN:
            row = [roll, rec["name"], "Data Restricted", "—", "—", "—", "—"]
        else:
            row = [
                roll, rec["name"], rec["status"],
                f"{rec.get('ai_confidence', 0)*100:.1f}%",
                override["new_status"] if override else "—",
                override["faculty_name"] if override else "—",
                override["comment"] if override else "—"
            ]
        ws.append(row)

    # Auto-width columns
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except AttributeError:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # SHA-256 footer
    ws.append([])
    digest = hashlib.sha256(str(records).encode()).hexdigest()
    ws.append([f"SHA-256: {digest}"])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
